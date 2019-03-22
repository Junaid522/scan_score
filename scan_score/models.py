
from django.db import models
from django.contrib.auth.models import User, AbstractUser
from django.db.models.signals import  post_save
from django.dispatch import  receiver
import openpyxl


# Create your models here.

from django.utils import timezone

class Document(models.Model):
    # description = models.CharField(max_length=255, blank=True)
    document = models.ImageField(upload_to='documents/')
    uploaded_at = models.DateTimeField(auto_now_add=True)




class AnswerFile(models.Model):

    test_file = models.FileField(blank=True)


    def read_file(self, val):

        test_type = TestType(name= "hahahahah")
        self.test_type.save()

    class Meta:

        verbose_name_plural = "AnswerFile"

# @receiver(post_save, sender = AnswerFile)
# def file_is_added(sender, instance, created, **kwargs):
#
#     if created:
#         instance.

def file_added(sender, **kwargs):

    if AnswerFile.objects.first():

        file_obj = AnswerFile.objects.first()
        file = file_obj.test_file
        print(file_obj.test_file)

        wb = openpyxl.load_workbook(file)
        worksheet = wb["ACTAnswers"]

        excel_test_data = list()
        # reading answer_file from file

        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_test_data.append(row_data)
        print(excel_test_data)


        TestType.objects.all().delete()
        TestName.objects.all().delete()
        TestSection.objects.all().delete()
        Question.objects.all().delete()

        for i in range(1, len(excel_test_data)):

            test_type = TestType(name=excel_test_data[i][0])
            test_type.save()
            test_name = TestName(name=excel_test_data[i][1], test_type_id= test_type)
            test_name.save()
            test_section = TestSection(section_number=excel_test_data[i][2], section_type=excel_test_data[i][3], test_name_id=test_name)
            test_section.save()
            question = Question(question_number=excel_test_data[i][4], answer_key=excel_test_data[i][5], test_section_id=test_section )
            question.save()

def percentile_file_added(sender, **kwargs):

    if PercentileFile.objects.first():

        file_obj = PercentileFile.objects.first()
        file = file_obj.percentile_file
        print(file_obj.percentile_file)
        wb = openpyxl.load_workbook(file)
        worksheet = wb["Sheet1"]

        excel_percentile_data = list()

        # reading from file

        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_percentile_data.append(row_data)
        print(excel_percentile_data)

        composite = int(excel_percentile_data[1][0])
        # print(composite)

        for i in range(1, len(excel_percentile_data)):

            percentile = Percentile(composite_score=composite, percentile=excel_percentile_data[i][1])
            percentile.save()
            composite -= 1






class TestType(models.Model):

    name = models.CharField(max_length=255, blank=True)


class TestName(models.Model):
    name = models.CharField(max_length=255, blank=True)
    test_type_id = models.ForeignKey(TestType, on_delete=models.CASCADE)


class TestSection(models.Model):

    section_number = models.CharField(max_length=255, blank=True)
    section_type = models.CharField(max_length=255, blank=True)
    test_name_id = models.ForeignKey(TestName, on_delete=models.CASCADE)


class Question(models.Model):

    question_number = models.CharField(max_length=255, blank=True)
    answer_key = models.CharField(max_length=255, blank=True)
    test_section_id = models.ForeignKey(TestSection, on_delete=models.CASCADE)


class PercentileFile(models.Model):

    percentile_file = models.FileField(blank=True)

    class Meta:
        verbose_name_plural = "PercentileFile"


class Percentile(models.Model):

    composite_score = models.CharField(max_length=255, blank=True)
    percentile = models.CharField(max_length=255, blank=True)

    def __str__(self):

        return (self.composite_score + " : "+ self.percentile)

# signal for AnswerFile
post_save.connect(file_added, sender=AnswerFile)
# signal for percentile
post_save.connect(percentile_file_added, sender=PercentileFile)